import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

df = pd.read_csv('yatra.csv')
total = df['Cost'].sum()
df.loc[len(df.index)] = ['Total', '', total]
df.to_excel('Yatra_Expense_Report.xlsx', index=False)

wb = load_workbook('Yatra_Expense_Report.xlsx')
ws = wb.active

header_font = Font(bold=True)
for cell in ws[1]:
    cell.font = header_font

total_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
for cell in ws[ws.max_row]:
    cell.fill = total_fill
    if cell.column == 1:
        cell.font = Font(bold=True)

wb.save('Yatra_Expense_Report.xlsx')
